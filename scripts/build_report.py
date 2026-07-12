# -*- coding: utf-8 -*-
"""
月度消费报告生成器（通用版）。

用法:
    python build_report.py <input.json> <output.xlsx>

输入 JSON 结构见 SKILL.md 的「输出报告」一节。核心字段:
    month              : 字符串，如 "2026-05"
    detail             : [[日期, 交易对方, 商品/说明, 大类, 金额(数字), 来源, 备注], ...]
                         大类必须是这 6 个之一: 固定支出/日常生活/健康医疗/孝亲/个人生活/社交
    receivables        : [[项目, 金额(数字), 状态/说明], ...]   代垫·待回收（应收，不计消费）
    excluded           : [[项目, 金额(数字), 原因], ...]        不计入消费的支出
    reconciliation     : [[信用卡交易, 金额(数字), 对应来源, 对账状态], ...]
    summary_notes      : [字符串, ...]   汇总页底部的口径说明（第一行通常是"口径说明："）
    optimize_headline  : 字符串         可优化页顶部的一句话结论
    optimize_actionable: [[项目, 现状, 年化支出(数字), 潜在可省/年(字符串), 建议], ...]
    optimize_toconfirm : [[待核查项, 年化支出(数字), 需要的信息, 若成立·可省/年], ...]
    optimize_notes     : [字符串, ...]

所有金额传数字（脚本负责货币格式）。SUMIF 按大类自动汇总，数值随明细联动。
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
CNY = '¥#,##0.00;(¥#,##0.00);-'
PCT = '0.0%'
CATS = ["固定支出", "日常生活", "健康医疗", "孝亲", "个人生活", "社交"]

title_font = Font(name=FONT, size=14, bold=True, color="FFFFFF")
h_font     = Font(name=FONT, size=10, bold=True, color="FFFFFF")
b_font     = Font(name=FONT, size=10)
bold_font  = Font(name=FONT, size=10, bold=True)
note_font  = Font(name=FONT, size=9, italic=True, color="666666")

title_fill = PatternFill("solid", fgColor="1F4E5F")
head_fill  = PatternFill("solid", fgColor="2E7D8A")
total_fill = PatternFill("solid", fgColor="DCE9EC")
excl_fill  = PatternFill("solid", fgColor="F4E1E1")
recv_fill  = PatternFill("solid", fgColor="FFF3D6")
save_fill  = PatternFill("solid", fgColor="E2F0E4")

thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = h_font; cell.fill = head_fill
        cell.alignment = center; cell.border = border


def write_row(ws, r, values, fills=None, money_cols=(), center_cols=(), font=None):
    font = font or b_font
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = font; cell.border = border
        if fills:
            cell.fill = fills
        if i in money_cols:
            cell.number_format = CNY; cell.alignment = right
        elif i in center_cols:
            cell.alignment = center
        else:
            cell.alignment = left


def build(data, out_path):
    month = data.get("month", "")
    wb = Workbook()

    # ---------- 消费明细 ----------
    det = wb.active
    det.title = "消费明细"
    det.merge_cells("A1:G1")
    det["A1"] = f"{month} · 消费明细（已去重、已按口径核定）"
    det["A1"].font = title_font; det["A1"].fill = title_fill; det["A1"].alignment = center
    det.row_dimensions[1].height = 26
    for i, h in enumerate(["日期", "交易对方", "商品/说明", "大类", "金额(元)", "来源", "备注"], 1):
        det.cell(row=2, column=i, value=h)
    style_header(det, 2, 7)
    r = 3
    for row in data.get("detail", []):
        write_row(det, r, row, money_cols=(5,), center_cols=(1, 4, 6))
        r += 1
    first_data, last_data = 3, r - 1
    det.cell(row=r, column=4, value="合计").font = bold_font
    det.cell(row=r, column=4).alignment = center
    tot = det.cell(row=r, column=5, value=f"=SUM(E{first_data}:E{last_data})")
    tot.font = bold_font; tot.number_format = CNY; tot.alignment = right
    for c in range(1, 8):
        det.cell(row=r, column=c).fill = total_fill; det.cell(row=r, column=c).border = border
    for i, w in enumerate([12, 14, 18, 10, 13, 8, 20], 1):
        det.column_dimensions[get_column_letter(i)].width = w

    # ---------- 消费汇总 ----------
    summ = wb.create_sheet("消费汇总", 0)
    summ.merge_cells("A1:C1")
    summ["A1"] = f"{month} 消费报告 · 分类汇总"
    summ["A1"].font = title_font; summ["A1"].fill = title_fill; summ["A1"].alignment = center
    summ.row_dimensions[1].height = 28
    for i, h in enumerate(["大类", "金额(元)", "占比"], 1):
        summ.cell(row=2, column=i, value=h)
    style_header(summ, 2, 3)
    r = 3; cat_start = r
    total_ref = cat_start + len(CATS)
    for cat in CATS:
        summ.cell(row=r, column=1, value=cat).font = b_font
        summ.cell(row=r, column=1).alignment = left
        f = (f"=SUMIF('消费明细'!$D${first_data}:$D${last_data},A{r},"
             f"'消费明细'!$E${first_data}:$E${last_data})")
        amt = summ.cell(row=r, column=2, value=f)
        amt.font = b_font; amt.number_format = CNY; amt.alignment = right
        pct = summ.cell(row=r, column=3, value=f"=IF($B${total_ref}=0,0,B{r}/$B${total_ref})")
        pct.font = b_font; pct.number_format = PCT; pct.alignment = right
        for c in range(1, 4):
            summ.cell(row=r, column=c).border = border
        r += 1
    summ.cell(row=r, column=1, value="合计").font = bold_font
    summ.cell(row=r, column=1).alignment = left
    t = summ.cell(row=r, column=2, value=f"=SUM(B{cat_start}:B{r-1})")
    t.font = bold_font; t.number_format = CNY; t.alignment = right
    p = summ.cell(row=r, column=3, value=f"=IF(B{r}=0,0,B{r}/B{r})")
    p.font = bold_font; p.number_format = PCT; p.alignment = right
    for c in range(1, 4):
        summ.cell(row=r, column=c).fill = total_fill; summ.cell(row=r, column=c).border = border
    nr = r + 2
    for line in data.get("summary_notes", []):
        summ.cell(row=nr, column=1, value=line).font = note_font
        summ.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=3)
        summ.cell(row=nr, column=1).alignment = left
        nr += 1
    for col, w in zip("ABC", [16, 15, 10]):
        summ.column_dimensions[col].width = w

    # ---------- 可优化消费项 ----------
    opt = wb.create_sheet("可优化消费项", 1)
    opt.merge_cells("A1:E1")
    opt["A1"] = "可优化消费项 · 省钱清单"
    opt["A1"].font = title_font; opt["A1"].fill = title_fill; opt["A1"].alignment = center
    opt.row_dimensions[1].height = 26
    r = 2
    headline = data.get("optimize_headline", "")
    if headline:
        opt.merge_cells(f"A{r}:E{r}")
        opt.cell(row=r, column=1, value=headline).font = bold_font
        opt.cell(row=r, column=1).alignment = left
        opt.row_dimensions[r].height = 20
        r += 2
    act = data.get("optimize_actionable", [])
    if act:
        opt.cell(row=r, column=1, value="一、可直接行动（订阅优化）").font = bold_font
        r += 1
        for i, h in enumerate(["项目", "现状", "年化支出", "潜在可省/年", "建议"], 1):
            opt.cell(row=r, column=i, value=h)
        style_header(opt, r, 5)
        r += 1
        for row in act:
            write_row(opt, r, row, fills=save_fill, money_cols=(3,))
            opt.row_dimensions[r].height = 46
            r += 1
        r += 1
    tc = data.get("optimize_toconfirm", [])
    if tc:
        opt.cell(row=r, column=1, value="二、待你确认后才能判定（单月数据不足，不硬猜）").font = bold_font
        r += 1
        for i, h in enumerate(["待核查项", "年化支出", "需要的信息", "若成立·可省/年"], 1):
            opt.cell(row=r, column=i, value=h)
        style_header(opt, r, 4)
        r += 1
        for row in tc:
            write_row(opt, r, row, fills=recv_fill, money_cols=(2,))
            opt.row_dimensions[r].height = 32
            r += 1
        r += 1
    for line in data.get("optimize_notes", []):
        opt.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        opt.cell(row=r, column=1, value=line).font = note_font
        opt.cell(row=r, column=1).alignment = left
        r += 1
    for col, w in zip("ABCDE", [22, 18, 14, 20, 42]):
        opt.column_dimensions[col].width = w

    # ---------- 剔除与对账 ----------
    ex = wb.create_sheet("剔除与对账")
    ex.merge_cells("A1:D1")
    ex["A1"] = "代垫待回收 · 剔除项 · 信用卡对账"
    ex["A1"].font = title_font; ex["A1"].fill = title_fill; ex["A1"].alignment = center
    ex.row_dimensions[1].height = 24
    r = 2
    recv = data.get("receivables", [])
    ex.cell(row=r, column=1, value="一、代垫·待回收（应收回，不计入消费）").font = bold_font
    r += 1
    for i, h in enumerate(["项目", "金额(元)", "状态/说明"], 1):
        ex.cell(row=r, column=i, value=h)
    style_header(ex, r, 3)
    r += 1
    recv_start = r
    for row in recv:
        write_row(ex, r, row, fills=recv_fill, money_cols=(2,))
        r += 1
    ex.cell(row=r, column=1, value="待回收合计").font = bold_font
    ex.cell(row=r, column=1).alignment = left
    if r > recv_start:
        c2 = ex.cell(row=r, column=2, value=f"=SUM(B{recv_start}:B{r-1})")
    else:
        c2 = ex.cell(row=r, column=2, value=0)
    c2.font = bold_font; c2.number_format = CNY; c2.alignment = right
    for c in range(1, 4):
        ex.cell(row=r, column=c).fill = recv_fill; ex.cell(row=r, column=c).border = border
    r += 2

    ex.cell(row=r, column=1, value="二、不计入消费的支出").font = bold_font
    r += 1
    for i, h in enumerate(["项目", "金额(元)", "原因"], 1):
        ex.cell(row=r, column=i, value=h)
    style_header(ex, r, 3)
    r += 1
    for row in data.get("excluded", []):
        write_row(ex, r, row, fills=excl_fill, money_cols=(2,))
        r += 1
    r += 1

    ex.cell(row=r, column=1, value="三、信用卡对账（与微信/支付宝重复，不重复计入）").font = bold_font
    r += 1
    for i, h in enumerate(["信用卡交易", "金额(元)", "对应来源", "对账状态"], 1):
        ex.cell(row=r, column=i, value=h)
    style_header(ex, r, 4)
    r += 1
    for row in data.get("reconciliation", []):
        write_row(ex, r, row, money_cols=(2,), center_cols=(3,))
        r += 1
    for col, w in zip("ABCD", [24, 14, 12, 30]):
        ex.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"saved -> {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("用法: python build_report.py <input.json> <output.xlsx>")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as fh:
        build(json.load(fh), sys.argv[2])
